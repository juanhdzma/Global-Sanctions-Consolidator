from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from os.path import join
from src.util.error import CustomError


def save_to_excel(df, filename):
    """Guarda el DataFrame en un archivo Excel y ajusta el ancho de las columnas"""
    try:
        full_path = join("./output_files/", filename)

        df.to_excel(full_path, index=False)

        wb = load_workbook(full_path)
        ws = wb.active

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(full_path)
        wb.close()
        return True
    except Exception:
        raise CustomError(f"❌ No se pudo guardar el archivo: {filename}")
